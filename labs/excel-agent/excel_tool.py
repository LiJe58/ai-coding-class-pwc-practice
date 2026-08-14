from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


DATA_PATH = Path(__file__).with_name("data") / "Excel_Copilot_anonymized.xlsx"
OUTPUT_COLUMNS = [
    "계정코드",
    "분류",
    "조서계정",
    "계정명",
    "계산된 당기(수정후)",
    "정산표 당기(수정후)(G)",
    "차이",
    "검증결과",
]


def get_pbc_mismatches(top_n: int) -> dict:
    """PBC 검증 결과에서 차이가 큰 불일치 항목을 반환합니다."""
    if isinstance(top_n, bool) or not isinstance(top_n, int) or not 1 <= top_n <= 20:
        raise ValueError("top_n은 1부터 20 사이의 정수여야 합니다.")
    if not DATA_PATH.is_file():
        raise FileNotFoundError("비식별 Excel 파일이 data 폴더에 없습니다.")

    workbook = load_workbook(DATA_PATH, read_only=True, data_only=True)
    try:
        sheet = workbook["PBC>>>"]
        rows = sheet.iter_rows(values_only=True)
        headers = next((list(row) for row in rows if "검증결과" in row and "차이" in row), None)
        if headers is None:
            raise ValueError("PBC 시트에서 검증 결과 열을 찾지 못했습니다.")
        records = [dict(zip(headers, row, strict=True)) for row in rows]
    finally:
        workbook.close()

    records = [row for row in records if row.get("검증결과") in {"일치", "불일치"}]
    counts = {status: sum(row["검증결과"] == status for row in records) for status in ("일치", "불일치")}
    mismatches = sorted(
        (row for row in records if row["검증결과"] == "불일치"),
        key=lambda row: abs(row.get("차이") or 0),
        reverse=True,
    )[:top_n]
    return {
        "counts": counts,
        "top_n": top_n,
        "items": [{column: row.get(column) for column in OUTPUT_COLUMNS} for row in mismatches],
        "requires_human_review": True,
    }


if __name__ == "__main__":
    print(get_pbc_mismatches(4))
